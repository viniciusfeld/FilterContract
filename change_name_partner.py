import pandas as pd
import openpyxl


class ChangeNamePartner():

    def __init__(self,):

        self.path_commission = r'C:\tributo\FilterContract\Planilhas\Planilha_-_Comissoes_-_2023-Copy.xlsx'
        self.path_partner = r'C:\tributo\FilterContract\Planilhas\AlteracaoNomeParceiros.xlsx'

    def change_name(self):

        partner_name = pd.read_excel(self.path_partner)

        wb = openpyxl.load_workbook(self.path_commission)

        sheet_name = wb.sheetnames

        for name in sheet_name:
            print(name)
            df_commission = pd.read_excel(self.path_commission, sheet_name=name)
            # print("comission", df_commission['PARCEIRO'])
            # print("partner name", partner_name['Nome antigo'])

            collumn_name = 'NOME' if name == 'base comissão' else 'PARCEIRO'

            merged_df = pd.merge(df_commission, partner_name, left_on=collumn_name, right_on='Nome antigo', how='left')

            merged_df.loc[merged_df['Nome antigo'] == merged_df[collumn_name], collumn_name] = merged_df['Nome novo']

            merged_df.drop(columns=['Nome antigo', 'Nome novo'], inplace=True)

            print(merged_df) 
 
            ws = wb[name] 

            collumn = 'B' if name == 'MARCO-022023' or name == 'ABRIL - 032023' or name == 'MAIO - 042023' else 'A'

            for idx, valor in enumerate(merged_df[collumn_name], start=2):
                ws[f'{collumn}{idx}'] = valor

                cont = 0
                for cell in ws[f"A{idx}:ZZ{idx}"][0]:
                    if name == 'SETEMBRO - 082023' or name == 'base comissão':
                        break

                    cont += 1
                    if cont == 8:
                        fees_value = cell.value

                    if cont == 9:
                        tax_value = cell.value

                    if cont == 10:
                        if cell.value == "TOTAL PAGO" or cell.value == "TOTAL GERAL":
                            break
                        if cell.value != None:
                            print("cell", cell.value)
                            if type(cell.value) == float:
                                try:
                                    without_tax = fees_value - tax_value
                                    percent_partner = cell.value / without_tax
                                    percent_partner = round(float(percent_partner) * 100, 2)
                                    print("percent_partner", percent_partner)
                                except:
                                    message = f"Erro: calculo diferente. Planilha: {name}/ Celula: {cell.value}/ Linha: {idx}\n"
                                    open(r'C:\tributo\FilterContract\error_contract.txt', "a").write(f"{str(message)}")
                                    print("deu erro na celula:", cell.value)
                                    break
                            
                            else:
                                try:
                                    percent_partner = cell.value.split("*")
                                    percent_partner = round(float(percent_partner[-1]) * 100, 2)
                                    print("percent partner", percent_partner)
                                except:
                                    message = f"Erro: calculo diferente. Planilha: {name}/ Celula: {cell.value}/ Linha: {idx}\n"
                                    open(r'C:\tributo\FilterContract\error_contract.txt', "a").write(f"{str(message)}")
                                    print("deu erro na celula:", cell.value)
                                    break
                        else:
                            break

                    if cont == 13:
                        if cell.value == 'cancelada' or cell.value == None:
                            break
                        else:
                            percent_partner_collumn = 'O'
                            ws[f'{percent_partner_collumn}{idx}'] = percent_partner
                            break

            wb.save(self.path_commission)

        wb.close()


    def percent_partner(self):
        partner_name = pd.read_excel(self.path_commission)


def main():
    change_name_partner = ChangeNamePartner()
    change_name_partner.change_name()

if __name__ == "__main__":
    main()